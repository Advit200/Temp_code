from tkinter import *
from tkinter import ttk
from ttkthemes import ThemedTk
import tkinter.messagebox as tmsg
import tkinter.filedialog as tkfd
import os
import shutil
import time
import pandas as pd
import openpyxl
import warnings
warnings.filterwarnings("ignore")
import AP322_docx_main

root=ThemedTk(theme="itft1")
root.geometry("1000x170")
root.maxsize(width=1000,height=515)
root.title("myConcerto | ADM Document Templatizer ( Powered by VIRTUOSO - An Intelligent CORE )")
root.wm_iconbitmap('icon.ico')
title_frame=Frame(root,bg='#9433c2')
title_frame.pack(side=TOP,fill=X)
Label(title_frame,text="Welcome to ADM Document Templatizer !",bg='#9337be',fg='white',font=("century schoolbook",20,"bold")).pack(side=LEFT,padx=130,pady=3)
phototitle=PhotoImage(file=r"myconcerto.png")
Label(title_frame,image=phototitle).pack(side=RIGHT,padx=3,pady=3)
global photosearch
photosearch=PhotoImage(file=r"search.png")
global photorefresh
photorefresh=PhotoImage(file=r"refresh.png")
global phototemplate
phototemplate=PhotoImage(file=r"template.png")
global photofilter
photofilter=PhotoImage(file=r"filter.png")
global photoclear
photoclear=PhotoImage(file=r"clear.png")

run=0
filterrun=0

def time_taken(func):
    def inner_function(*args,**kwargs):
        start=time.time()
        func(*args,**kwargs)
        end=time.time()
        print('Time taken is : {:.2f} sec'.format(end-start))
    return inner_function

def output_function():
    global output_entry
    global output_dir
    output_dir=tkfd.askdirectory()
    output_entry.delete(0,'end')
    output_entry.insert(0,output_dir)
    output_dir=output_entry.get()

    def output_folder_creation(output_dir):
        folders_list=['Doc_Template','Images','Spell_check','Final_report','Detailed_report']
        try:
            for i in folders_list:
                op_path=os.path.join(output_dir,'Output_Files',i)
                os.makedirs(op_path)
        except FileExistsError as e:
            ans=tmsg.askyesno("ERROR", "Selected directory has existing Output_Files folder.\n\n Do you want to override the folder ? ")
            if ans:
                del_path=os.path.join(output_dir,'Output_Files')
                shutil.rmtree(del_path)
                for i in folders_list:
                    op_path=os.path.join(output_dir,'Output_Files',i)
                    os.makedirs(op_path)
            else:
                output_function()
    if output_dir:
        output_folder_creation(output_dir)
    else:
        ans=tmsg.askyesno("ERROR", "Select Output Folder.")
        if ans:
            output_function()
        else:
            reset_function()
        
def output_frame_function():
    global output_frame
    global output_entry
    output_frame=ttk.Frame(root)
    output_frame.pack(side=TOP,fill=X)
    ttk.Label(output_frame,text='Select the Output Folder',font=('century schoolbook',11,"bold")).pack(side=LEFT,padx=80)
    ttk.Button(output_frame,text="Browse",command=output_function).pack(side=LEFT)
    output_entry=ttk.Entry(output_frame,width=80)
    output_entry.pack(side=LEFT,padx=22,fill=X)  

def filter_frame_function():
    global admcombobox
    global doccombobox
    global filter_frame
    global photofilter
    global admfiltervar
    global docfiltervar
    filter_frame=ttk.Frame(root)
    filter_frame.pack(side=TOP,fill=X)
    ttk.Label(filter_frame,text="Filter by ADM Number",font=('century schoolbook',11)).pack(side=LEFT,padx=(80,10),pady=20)
    admfiltervar=StringVar()
    admcombobox=ttk.Combobox(filter_frame,textvariable=admfiltervar,values=['All ADM Documents','BP310','BP315','AP322','TE586','TR435'])
    admcombobox.pack(side=LEFT,padx=(15,0),pady=20)
    admcombobox.current(0)
    
    ttk.Label(filter_frame,text="Filter by Document format",font=('century schoolbook',11)).pack(side=LEFT,padx=(50,10),pady=20)
    docfiltervar=StringVar()
    doccombobox=ttk.Combobox(filter_frame,textvariable=docfiltervar,values=['All Formats','docx','xlsx','pptx'])
    doccombobox.pack(side=LEFT,padx=17,pady=20)
    doccombobox.current(0)
    ttk.Button(filter_frame,text="Filter",image=photofilter,compound=LEFT,command=custom_filter_function).pack(side=LEFT,padx=10)

def search_function():
    global final_filter_file
    global final_input_list
    global filterrun
    global temp_input_files
    global user_search
    global listbox
    final_input_list=[]
    if filterrun==0:
        final_filter_file=temp_input_files
        for file in final_filter_file:
            if user_search.get().lower() in file.lower():
                final_input_list.append(file)
        print(final_input_list)
        listbox.delete(0,END)
        [listbox.insert(END,i) for i in final_input_list]
    else:
        for file in final_filter_file:
            if user_search.get().lower() in file.lower():
                final_input_list.append(file)
        print(final_input_list)
        listbox.delete(0,END)
        [listbox.insert(END,i) for i in final_input_list]

def clear_filter_function():
    global admcombobox
    global doccombobox
    global search_entry
    global listbox
    global temp_input_files
    admcombobox.set("All ADM Documents")
    doccombobox.set("All Formats")
    search_entry.delete(0,'end')
    listbox.delete(0,END)
    [listbox.insert(END,i) for i in temp_input_files]

def search_frame_function():
    global search_frame
    global search_entry
    global photosearch
    global photorefresh
    global user_search
    search_frame=ttk.Frame(root)
    search_frame.pack(side=TOP,fill=X)
    user_search=StringVar()
    ttk.Label(search_frame,text="Search with File Name",font=('century schoolbook',11)).pack(side=LEFT,padx=(80,0),pady=(0,5))
    search_entry=Entry(search_frame,textvariable=user_search,width=48,bd=3,font=("times new roman",11),relief=RAISED)
    search_entry.pack(anchor=NE,side=LEFT,padx=(10,15),pady=(7,15))
    ttk.Button(search_frame,image=photosearch,compound=LEFT,text="Search",command=search_function).pack(anchor=NE,side=LEFT,padx=(5,0),pady=(0,5))
    ttk.Button(search_frame,image=photoclear,text="Clear Filters",compound=LEFT,command=clear_filter_function).pack(anchor=NE,side=LEFT,padx=(10,0),pady=(3,5))
    ttk.Button(search_frame,image=photorefresh,compound=LEFT,text="Reset Window",command=reset_function).pack(anchor=NE,side=LEFT,padx=(10,10),pady=(0,5))

def link_creation(report_name,col):
    '''This function generates the final excel report with the link of detailed report as one column'''

    global output_dir
    final_report_path=os.path.join(output_dir,'Output_Files','Final_report',report_name)
    workbook=openpyxl.load_workbook(filename=final_report_path + '.xlsx')
    ws_object=workbook.active
    ws_object[col+str(1)] ='Detailed Report Link' # L1 means column L and row 1 (needs to be changed accordingly when more columns are added beforehand.)
    
    for i in range(2,ws_object.max_row+1): # skipping row 1 as it is the heading.
        file_link=os.path.splitext(ws_object.cell(row=i,column=1).value)[0]+'__DetailedReport'+'.xlsx'
        detail_report_link=os.path.join(output_dir,'Output_Files','Detailed_report',file_link)
        row=col+str(i)
        ws_object[row].hyperlink=detail_report_link
        ws_object[row].value='Report'
        ws_object[row].style='Hyperlink'
    
    temp=final_report_path + '.xlsx'
    workbook.save(filename=temp) # To update the changes we did above to the same excel file.

@time_taken
def report_generation_function():
    global listbox
    global selected_files_list
    global input_dir
    global output_dir
    global statusvar
    global sbar

    statusvar.set('Kernal Busy . Generating Report...')
    sbar.update()

    all_items=listbox.get(0,END)
    selected_indices=listbox.curselection()
    selected_files_list=[all_items[i] for i in selected_indices]

    if len(selected_files_list) == 0:
        tmsg.showinfo("ERROR","No Files selected .\n\n Please select a file !")

    else:
        try:
            ap322_tem_found , _ = AP322_docx_main.ap322_template_checking(input_dir,output_dir)
            if ap322_tem_found :            
                ap322_result = [AP322_docx_main.running_ap322_class_methods(input_dir,file_name,output_dir) for file_name in selected_files_list if "AP322" in file_name and os.path.splitext(file_name)[1]==".docx"]
                final_dict=ap322_result[-1]
                
                final_report_df=pd.DataFrame(final_dict)
                print(final_report_df)
                final_report_df=final_report_df.style.applymap(lambda x: "background-color: #ffb366" if x=='Not Matching' and isinstance(x,str)  else "background-color: #ffcc99" if isinstance(x,int) and 0<x<=3 else "background-color: #ffb366" if isinstance(x,int) and 3<x<=6 else "background-color: #ff8c1a" if isinstance(x,int) and 6<x<=10 else "background-color: #ff8000" if isinstance(x,int) and 10<x<=15 else "background-color: #ff5c33" if isinstance(x,int) and x>15 else "background-color: white")
                FinalReport_path=os.path.join(output_dir,'Output_Files','Final_report','AP322_FinalReport')
                final_report_df.to_excel(FinalReport_path + '.xlsx',index=False,sheet_name='FinalConsolidatedReport')
                link_creation(report_name='AP322_FinalReport',col='L')
 
            tmsg.showinfo("ADM Document Templatizer","Report Generation Completed ! \n\n Thanks for using the Tool.")
            statusvar.set('Ready Now...')

        except NameError as e:
            print(e)
            tmsg.showinfo("ERROR","Select Output Folder !")
        except Exception as e:
            print(e)
            tmsg.showinfo("ERROR",'No Template Found in the selected Input Folder.')
        
def listbox_frame_function():
    global listbox_frame
    global input_dir
    global listbox
    global phototemplate
    global final_input_list
    listbox_frame=ttk.Frame(root)
    listbox_frame.pack(side=TOP,fill=X)
    scrollbar=ttk.Scrollbar(listbox_frame)
    scrollbar.pack(side=RIGHT,fill=Y,padx=(0,300))
    listbox=Listbox(listbox_frame,height=15,width=120,selectmode=MULTIPLE,bd=2,yscrollcommand=scrollbar.set)
    listbox.pack(side=LEFT,anchor=NW,padx=(80,0),pady=(5,0))
    scrollbar.config(command=listbox.yview)
    ttk.Button(listbox_frame,text="Select All Files",command=lambda:listbox.select_set(0,END)).place(x=730,y=25)
    ttk.Button(listbox_frame,text="Clear all selections",command=lambda:listbox.select_clear(0,END)).place(x=730,y=70)
    ttk.Button(listbox_frame,image=phototemplate,compound=LEFT,text="Generate Report",command=report_generation_function).place(x=730,y=170)
    [listbox.insert(END,i) for i in final_input_list]

def custom_filter_function():
    global doccombobox_filter_result
    global admcombobox_filter_result
    global final_filter_file
    global doccombobox
    global filterrun
    global temp_input_files
    global admcombobox_filter_result
    global listbox
    final_filter_file=[]
    doccombobox_filter_result=doccombobox.get()
    admcombobox_filter_result=admcombobox.get()
    print(doccombobox_filter_result)
    print(admcombobox_filter_result)
    for no_of_run in range(1):
        if admcombobox_filter_result =="All ADM Documents" and doccombobox_filter_result=="All Formats":
            final_filter_file=temp_input_files
        elif admcombobox_filter_result=="All ADM Documents" and doccombobox_filter_result=="docx":
            for file in temp_input_files:
                if os.path.splitext(file)[1]==".docx":
                    final_filter_file.append(file)
        elif admcombobox_filter_result=="All ADM Documents" and doccombobox_filter_result=="xlsx":
            for file in temp_input_files:
                if os.path.splitext(file)[1]==".xlsx":
                    final_filter_file.append(file)
        elif admcombobox_filter_result=="All ADM Documents" and doccombobox_filter_result=="pptx":
            for file in temp_input_files:
                if os.path.splitext(file)[1]==".pptx":
                    final_filter_file.append(file)                        

        elif admcombobox_filter_result=="BP310" and doccombobox_filter_result=="docx":
            for file in temp_input_files:
                if "BP310" in file and os.path.splitext(file)[1]==".docx":
                    final_filter_file.append(file)
        elif admcombobox_filter_result=="BP310" and doccombobox_filter_result=="xlsx":
            for file in temp_input_files:
                if "BP310" in file and os.path.splitext(file)[1]==".xlsx":
                    final_filter_file.append(file)
        elif admcombobox_filter_result=="BP310" and doccombobox_filter_result=="pptx":
            for file in temp_input_files:
                if "BP310" in file and os.path.splitext(file)[1]==".pptx":
                    final_filter_file.append(file)
        elif admcombobox_filter_result=="BP310" and doccombobox_filter_result=="All Formats":
            for file in temp_input_files:
                if "BP310" in file:
                    final_filter_file.append(file)

        elif admcombobox_filter_result=="BP315" and doccombobox_filter_result=="docx":
            for file in temp_input_files:
                if "BP315" in file and os.path.splitext(file)[1]==".docx":
                    final_filter_file.append(file)
        elif admcombobox_filter_result=="BP315" and doccombobox_filter_result=="xlsx":
            for file in temp_input_files:
                if "BP315" in file and os.path.splitext(file)[1]==".xlsx":
                    final_filter_file.append(file)
        elif admcombobox_filter_result=="BP315" and doccombobox_filter_result=="pptx":
            for file in temp_input_files:
                if "BP315" in file and os.path.splitext(file)[1]==".pptx":
                    final_filter_file.append(file)
        elif admcombobox_filter_result=="BP315" and doccombobox_filter_result=="All Formats":
            for file in temp_input_files:
                if "BP315" in file:
                    final_filter_file.append(file)    

        elif admcombobox_filter_result=="AP322" and doccombobox_filter_result=="docx":
            for file in temp_input_files:
                if "AP322" in file and os.path.splitext(file)[1]==".docx":
                    final_filter_file.append(file)
        elif admcombobox_filter_result=="AP322" and doccombobox_filter_result=="xlsx":
            for file in temp_input_files:
                if "AP322" in file and os.path.splitext(file)[1]==".xlsx":
                    final_filter_file.append(file)
        elif admcombobox_filter_result=="AP322" and doccombobox_filter_result=="pptx":
            for file in temp_input_files:
                if "AP322" in file and os.path.splitext(file)[1]==".pptx":
                    final_filter_file.append(file)
        elif admcombobox_filter_result=="AP322" and doccombobox_filter_result=="All Formats":
            for file in temp_input_files:
                if "AP322" in file:
                    final_filter_file.append(file)

        elif admcombobox_filter_result=="TE586" and doccombobox_filter_result=="docx":
            for file in temp_input_files:
                if "TE586" in file and os.path.splitext(file)[1]==".docx":
                    final_filter_file.append(file)
        elif admcombobox_filter_result=="TE586" and doccombobox_filter_result=="xlsx":
            for file in temp_input_files:
                if "TE586" in file and os.path.splitext(file)[1]==".xlsx":
                    final_filter_file.append(file)
        elif admcombobox_filter_result=="TE586" and doccombobox_filter_result=="pptx":
            for file in temp_input_files:
                if "TE586" in file and os.path.splitext(file)[1]==".pptx":
                    final_filter_file.append(file)
        elif admcombobox_filter_result=="TE586" and doccombobox_filter_result=="All Formats":
            for file in temp_input_files:
                if "TE586" in file:
                    final_filter_file.append(file)            

        elif admcombobox_filter_result=="TR435" and doccombobox_filter_result=="docx":
            for file in temp_input_files:
                if "TR435" in file and os.path.splitext(file)[1]==".docx":
                    final_filter_file.append(file)
        elif admcombobox_filter_result=="TR435" and doccombobox_filter_result=="xlsx":
            for file in temp_input_files:
                if "TR435" in file and os.path.splitext(file)[1]==".xlsx":
                    final_filter_file.append(file)
        elif admcombobox_filter_result=="TR435" and doccombobox_filter_result=="pptx":
            for file in temp_input_files:
                if "TR435" in file and os.path.splitext(file)[1]==".pptx":
                    final_filter_file.append(file)
        elif admcombobox_filter_result=="TR435" and doccombobox_filter_result=="All Formats":
            for file in temp_input_files:
                if "TR435" in file:
                    final_filter_file.append(file)

    filterrun+=1                
    print(final_filter_file)
    listbox.delete(0,END)
    [listbox.insert(END,i) for i in final_filter_file]

def statusbar_frame_function():
    global statusbar_frame
    global statusvar
    global sbar
    statusbar_frame=ttk.Frame(root)
    statusbar_frame.pack(side=TOP,fill='both')
    statusvar=StringVar()
    statusvar.set("Ready...")
    sbar=Label(statusbar_frame,bg="#9433c2",fg="white",textvariable=statusvar,font=('century schoolbook',9,"bold"),relief=RAISED,anchor=W)
    sbar.pack(side=BOTTOM,fill=X,pady=(3,0))

def reset_function():
    global output_entry
    global input_entry
    global listbox
    global admcombobox
    global doccombobox
    global search_entry
    global input_dir
    global output_dir
    input_entry.delete(0,'end')
    output_entry.delete(0,'end')
    search_entry.delete(0,'end')
    listbox.delete(0,END)
    admcombobox.delete(0,END)
    doccombobox.delete(0,END)
    input_dir=None
    output_dir=None
    
def browse_refresh():
    global output_frame
    global input_entry
    global search_frame
    global listbox_frame
    global filter_frame
    global statusbar_frame
    input_entry.delete(0,'end')
    output_frame.destroy()
    filter_frame.destroy()
    search_frame.destroy()
    listbox_frame.destroy()
    statusbar_frame.destroy()   
          
def input_frame_function():
    global input_dir
    global run
    global temp_input_files
    global final_input_list
    # this if-else based on run variable is written to maintain the functionality when user clicks the browse button second time , then the window should refresh and so the input data.
    if run==0:
        try:
            input_dir=tkfd.askdirectory()
            temp_input_files=[k for i,j,k in os.walk(r"{}".format(input_dir))][0]
            final_input_list=temp_input_files
            temp_frame.destroy()
            run+=1
            input_entry.insert(0,input_dir)
            root.geometry("1000x515")
            output_frame_function()
            filter_frame_function()
            search_frame_function()        
            listbox_frame_function()
            statusbar_frame_function()
        except IndexError as e:
            print(e)
            root.geometry("1000x100")
        
    elif run != 0:
        try:
            temp_frame.destroy()
            browse_refresh()
            input_dir=tkfd.askdirectory()
            temp_input_files=[k for i,j,k in os.walk(r"{}".format(input_dir))][0]
            final_input_list=temp_input_files
            input_entry.insert(0,input_dir)
            root.geometry("1000x515")
            output_frame_function()
            filter_frame_function()
            search_frame_function()        
            listbox_frame_function()
            statusbar_frame_function()
        except IndexError as e:
            print(e)
            root.geometry("1000x100")
              
input_frame=ttk.Frame(root)
input_frame.pack(side=TOP,fill=X)
ttk.Label(input_frame,text='Select the Input Folder',font=('century schoolbook',11,"bold")).pack(side=LEFT,padx=80,pady=5)
ttk.Button(input_frame,text="Browse",command=input_frame_function).pack(side=LEFT,padx=11,pady=5)
input_entry=ttk.Entry(input_frame,width=80)
input_entry.pack(side=LEFT,padx=11,pady=5,fill=X)
# temp_frame to cover up the blank spaces on the screen.
temp_frame=ttk.Frame(root)
temp_frame.pack(side=TOP,fill='both')
ttk.Label(temp_frame).pack(side=TOP,fill='both',pady=30)


root.mainloop()


    
